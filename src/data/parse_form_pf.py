"""
SEC Form PF Parser

Extracts hedge fund industry metrics from Form PF Supporting Data Excel files.
The latest file (form_pf_2025Q1.xlsx) contains the full historical time series
from 2013Q1 onward across 141 sheets.

Produces 12 processed CSVs covering fund counts, GAV/NAV, borrowing, derivatives,
concentration, strategy allocation, leverage distributions, notional exposures,
liquidity, fair value hierarchy, geography, and sector data.
"""

import os
import pandas as pd
import numpy as np
import openpyxl

DATA_DIR = os.path.join(os.path.dirname(__file__), '..', '..', 'data', 'raw', 'form_pf')
OUTPUT_DIR = os.path.join(os.path.dirname(__file__), '..', '..', 'data', 'processed')


def _find_latest_excel(data_dir):
    """Find the most recent Form PF Excel file (contains full history)."""
    files = sorted([f for f in os.listdir(data_dir) if f.endswith('.xlsx')])
    if not files:
        raise FileNotFoundError(f"No Form PF Excel files in {data_dir}")
    return os.path.join(data_dir, files[-1])


def parse_simple_table(filepath, sheet_name, label_col='label'):
    """Parse a sheet with one label column and quarterly/monthly date columns.

    Layout: Row 0 = headers (label_name, date1, date2, ...)
            Rows 1-N = data
            Last row = 'Return to Notes' (sentinel)
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb[sheet_name]

    rows = []
    for row in ws.iter_rows(values_only=True):
        if row[0] and str(row[0]).strip() == 'Return to Notes':
            break
        rows.append(list(row))
    wb.close()

    if len(rows) < 2:
        return pd.DataFrame()

    header = rows[0]
    header[0] = label_col
    data = rows[1:]

    df = pd.DataFrame(data, columns=header)
    # Drop completely empty rows
    df = df.dropna(how='all', subset=header[1:])
    return df


def parse_two_level_table(filepath, sheet_name, col_names=None):
    """Parse a sheet with two label columns and date columns.

    Layout: Row 0 = headers (col1_name, col2_name, date1, date2, ...)
            Rows 1-N = data
            Last row = 'Return to Notes' (sentinel)
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb[sheet_name]

    rows = []
    for row in ws.iter_rows(values_only=True):
        if row[0] and str(row[0]).strip() == 'Return to Notes':
            break
        rows.append(list(row))
    wb.close()

    if len(rows) < 2:
        return pd.DataFrame()

    header = rows[0]
    if col_names:
        header[0] = col_names[0]
        header[1] = col_names[1]
    data = rows[1:]

    df = pd.DataFrame(data, columns=header)
    df = df.dropna(how='all', subset=header[2:])
    return df


def _melt_quarterly(df, label_col='label', value_name='value'):
    """Melt a wide quarterly table to long format with quarter column."""
    date_cols = [c for c in df.columns if c != label_col and str(c).startswith('20')]
    melted = df.melt(id_vars=[label_col], value_vars=date_cols,
                     var_name='quarter', value_name=value_name)
    melted[value_name] = pd.to_numeric(melted[value_name], errors='coerce')
    return melted


def _melt_monthly(df, label_col='label', value_name='value'):
    """Melt a wide monthly table to long format with month column."""
    date_cols = [c for c in df.columns if c != label_col and str(c).startswith('20')]
    melted = df.melt(id_vars=[label_col], value_vars=date_cols,
                     var_name='month', value_name=value_name)
    melted[value_name] = pd.to_numeric(melted[value_name], errors='coerce')
    return melted


def _melt_two_level(df, col1, col2, value_name='value', time_col='month'):
    """Melt a two-level wide table to long format."""
    date_cols = [c for c in df.columns if c not in [col1, col2] and str(c).startswith('20')]
    melted = df.melt(id_vars=[col1, col2], value_vars=date_cols,
                     var_name=time_col, value_name=value_name)
    melted[value_name] = pd.to_numeric(melted[value_name], errors='coerce')
    return melted


# ---------------------------------------------------------------------------
# Individual parsers for each output CSV
# ---------------------------------------------------------------------------

def parse_fund_counts(filepath):
    """Tab.1.1-1.4: Number of funds by type (quarterly)."""
    dfs = []
    for tab in ['Tab.1.1', 'Tab.1.2', 'Tab.1.3', 'Tab.1.4']:
        try:
            df = parse_simple_table(filepath, tab, label_col='fund_type')
            df['table'] = tab
            dfs.append(df)
        except Exception:
            continue
    if not dfs:
        return pd.DataFrame()

    # Tab.1.1 is the primary fund count table
    df = dfs[0]
    return _melt_quarterly(df, label_col='fund_type', value_name='count')


def parse_gav_nav(filepath):
    """Tab.2.1 (GAV), Tab.2.3 (NAV), Tab.2.5 (other) — quarterly, billions USD."""
    results = []

    # Tab.2.1 = GAV by fund type
    df = parse_simple_table(filepath, 'Tab.2.1', label_col='fund_type')
    melted = _melt_quarterly(df, label_col='fund_type', value_name='gav')
    results.append(melted)

    # Tab.2.3 = NAV by fund type
    df = parse_simple_table(filepath, 'Tab.2.3', label_col='fund_type')
    melted = _melt_quarterly(df, label_col='fund_type', value_name='nav')
    results.append(melted)

    # Merge GAV and NAV
    merged = results[0].merge(results[1], on=['fund_type', 'quarter'], how='outer')

    # Compute derived metrics
    merged['gav_nav_ratio'] = merged['gav'] / merged['nav']

    return merged


def parse_borrowing(filepath):
    """Tab.2.9, 2.13 (quarterly), Tab.8.27 (monthly detail), Tab.8.34 (quarterly creditor)."""
    parts = []

    # Tab.2.9: Borrowing as % of GAV (quarterly)
    df = parse_simple_table(filepath, 'Tab.2.9', label_col='fund_type')
    melted = _melt_quarterly(df, label_col='fund_type', value_name='borrowing_pct_gav')
    melted['source'] = 'Tab.2.9'
    parts.append(melted)

    # Tab.8.27: Borrowing detail by type (monthly) — two-level
    df = parse_two_level_table(filepath, 'Tab.8.27', col_names=['type', 'subtype'])
    melted = _melt_two_level(df, 'type', 'subtype', value_name='amount_bn', time_col='month')
    melted['source'] = 'Tab.8.27'

    # Tab.8.34: Borrowing by creditor type (monthly) — two-level
    try:
        df34 = parse_two_level_table(filepath, 'Tab.8.34', col_names=['creditor_type', 'subtype'])
        melted34 = _melt_two_level(df34, 'creditor_type', 'subtype',
                                   value_name='amount_bn', time_col='month')
        melted34['source'] = 'Tab.8.34'
    except Exception:
        melted34 = pd.DataFrame()

    return {
        'quarterly_pct': parts[0] if parts else pd.DataFrame(),
        'monthly_detail': melted,
        'creditor_type': melted34,
    }


def parse_derivatives(filepath):
    """Tab.5.1 (derivative value), Tab.5.3 (derivatives as % of NAV) — quarterly."""
    # Tab.5.1
    df1 = parse_simple_table(filepath, 'Tab.5.1', label_col='fund_type')
    melted1 = _melt_quarterly(df1, label_col='fund_type', value_name='derivative_value')

    # Tab.5.3
    df3 = parse_simple_table(filepath, 'Tab.5.3', label_col='fund_type')
    melted3 = _melt_quarterly(df3, label_col='fund_type', value_name='derivative_pct_nav')

    merged = melted1.merge(melted3, on=['fund_type', 'quarter'], how='outer')
    return merged


def parse_concentration(filepath):
    """Tab.6.3-6.6: Top N fund share of NAV, GAV, borrowing, derivatives — quarterly."""
    tabs = {
        'Tab.6.3': 'nav_share',
        'Tab.6.4': 'gav_share',
        'Tab.6.5': 'borrowing_share',
        'Tab.6.6': 'derivative_share',
    }
    results = []
    for tab, metric in tabs.items():
        try:
            df = parse_simple_table(filepath, tab, label_col='top_n')
            melted = _melt_quarterly(df, label_col='top_n', value_name=metric)
            melted['table'] = tab
            results.append(melted)
        except Exception:
            continue

    if not results:
        return pd.DataFrame()

    # Merge all concentration metrics
    merged = results[0]
    for r in results[1:]:
        merged = merged.merge(r.drop(columns='table', errors='ignore'),
                              on=['top_n', 'quarter'], how='outer')
    return merged


def parse_strategy(filepath):
    """Tab.8.7-8.15: Strategy allocation — GAV, NAV, borrowing, derivatives by strategy."""
    parts = []

    # Tab.8.9 = QHF GAV by strategy (quarterly)
    try:
        df = parse_simple_table(filepath, 'Tab.8.9', label_col='strategy')
        melted = _melt_quarterly(df, label_col='strategy', value_name='gav')
        melted['metric'] = 'gav'
        parts.append(melted)
    except Exception:
        pass

    # Tab.8.10 = QHF NAV by strategy (quarterly)
    try:
        df = parse_simple_table(filepath, 'Tab.8.10', label_col='strategy')
        melted = _melt_quarterly(df, label_col='strategy', value_name='nav')
        melted['metric'] = 'nav'
        parts.append(melted)
    except Exception:
        pass

    # Tab.8.14 = QHF borrowing by strategy (quarterly)
    try:
        df = parse_simple_table(filepath, 'Tab.8.14', label_col='strategy')
        melted = _melt_quarterly(df, label_col='strategy', value_name='borrowing')
        melted['metric'] = 'borrowing'
        parts.append(melted)
    except Exception:
        pass

    if not parts:
        return pd.DataFrame()

    # Merge GAV and NAV
    merged = parts[0][['strategy', 'quarter', 'gav']].copy() if len(parts) > 0 else pd.DataFrame()
    if len(parts) > 1:
        nav_df = parts[1][['strategy', 'quarter', 'nav']]
        merged = merged.merge(nav_df, on=['strategy', 'quarter'], how='outer')
    if len(parts) > 2:
        borrow_df = parts[2][['strategy', 'quarter', 'borrowing']]
        merged = merged.merge(borrow_df, on=['strategy', 'quarter'], how='outer')

    return merged


def parse_leverage_distribution(filepath):
    """Tab.8.1-8.6: GNE/LNE/SNE ratio distributions (monthly, fund counts per bucket)."""
    tabs = {
        'Tab.8.1': 'GNE',
        'Tab.8.2': 'LNE',
        'Tab.8.3': 'SNE',
        'Tab.8.4': 'GNE_excl_IRD',
        'Tab.8.5': 'LNE_excl_IRD',
        'Tab.8.6': 'SNE_excl_IRD',
    }
    results = []
    for tab, exposure_type in tabs.items():
        try:
            df = parse_two_level_table(filepath, tab, col_names=['exposure', 'ratio_bucket'])
            melted = _melt_two_level(df, 'exposure', 'ratio_bucket',
                                     value_name='fund_count', time_col='month')
            melted['exposure_type'] = exposure_type
            results.append(melted)
        except Exception:
            continue

    if not results:
        return pd.DataFrame()
    return pd.concat(results, ignore_index=True)


def parse_notional(filepath):
    """Tab.8.16 (long notional), Tab.8.17 (short notional) — monthly, by investment type."""
    # Tab.8.16 = Long notional
    df_long = parse_simple_table(filepath, 'Tab.8.16', label_col='investment_type')
    long_melted = _melt_monthly(df_long, label_col='investment_type', value_name='long_notional')

    # Tab.8.17 = Short notional
    df_short = parse_simple_table(filepath, 'Tab.8.17', label_col='investment_type')
    short_melted = _melt_monthly(df_short, label_col='investment_type', value_name='short_notional')

    merged = long_melted.merge(short_melted, on=['investment_type', 'month'], how='outer')
    merged['net_exposure'] = merged['long_notional'] - merged['short_notional']

    return merged


def parse_liquidity(filepath):
    """Tab.8.22 (investor liquidity), Tab.8.23 (portfolio liquidity), Tab.8.33 (financing)."""
    tabs = {
        'Tab.8.22': 'investor_liquidity',
        'Tab.8.23': 'portfolio_liquidity',
        'Tab.8.33': 'financing_liquidity',
    }
    results = []
    for tab, liq_type in tabs.items():
        try:
            df = parse_simple_table(filepath, tab, label_col='period')
            melted = _melt_quarterly(df, label_col='period', value_name='cumulative_pct')
            melted['liquidity_type'] = liq_type
            results.append(melted)
        except Exception:
            continue

    if not results:
        return pd.DataFrame()
    return pd.concat(results, ignore_index=True)


def parse_fair_value(filepath):
    """Tab.2.14-2.23: Fair value hierarchy (Level 1/2/3 assets & liabilities)."""
    results = []
    for i in range(14, 24):
        tab = f'Tab.2.{i}'
        try:
            df = parse_simple_table(filepath, tab, label_col='category')
            melted = _melt_quarterly(df, label_col='category', value_name='amount')
            melted['table'] = tab
            results.append(melted)
        except Exception:
            continue

    if not results:
        return pd.DataFrame()
    return pd.concat(results, ignore_index=True)


def parse_geography(filepath):
    """Tab.3.1-3.2: Geographic distribution."""
    results = []
    for tab in ['Tab.3.1', 'Tab.3.2']:
        try:
            df = parse_simple_table(filepath, tab, label_col='region')
            melted = _melt_quarterly(df, label_col='region', value_name='amount')
            melted['table'] = tab
            results.append(melted)
        except Exception:
            continue

    if not results:
        return pd.DataFrame()
    return pd.concat(results, ignore_index=True)


def parse_sector(filepath):
    """Tab.10.1-10.6: Industry/sector allocation (annual, Q4 only)."""
    results = []
    for i in range(1, 7):
        tab = f'Tab.10.{i}'
        try:
            df = parse_simple_table(filepath, tab, label_col='sector')
            melted = _melt_quarterly(df, label_col='sector', value_name='amount')
            melted['table'] = tab
            results.append(melted)
        except Exception:
            continue

    if not results:
        return pd.DataFrame()
    return pd.concat(results, ignore_index=True)


# ---------------------------------------------------------------------------
# Derived metrics
# ---------------------------------------------------------------------------

def compute_form_pf_metrics(gav_nav_df, concentration_df, notional_df, liquidity_df, strategy_df):
    """Compute derived metrics from parsed Form PF data."""
    metrics = {}

    # --- GAV/NAV ratio for hedge funds ---
    hf = gav_nav_df[gav_nav_df['fund_type'] == 'Hedge Fund'].copy()
    if not hf.empty:
        metrics['hf_gav_nav'] = hf[['quarter', 'gav', 'nav', 'gav_nav_ratio']].copy()
        metrics['hf_gav_nav']['gav_qoq'] = hf['gav'].pct_change()
        metrics['hf_gav_nav']['nav_qoq'] = hf['nav'].pct_change()

    # --- Concentration trend (top-10 NAV share slope) ---
    if not concentration_df.empty and 'nav_share' in concentration_df.columns:
        top10 = concentration_df[concentration_df['top_n'] == 'Top 10'].copy()
        if not top10.empty:
            metrics['concentration_top10'] = top10[['quarter', 'nav_share']].copy()

    # --- Net notional exposure by type (latest quarter) ---
    if not notional_df.empty:
        latest_month = notional_df['month'].max()
        latest = notional_df[notional_df['month'] == latest_month].copy()
        latest = latest.sort_values('net_exposure', ascending=False)
        metrics['latest_notional'] = latest

    # --- Strategy HHI (quarterly) ---
    if not strategy_df.empty and 'nav' in strategy_df.columns:
        # Exclude 'Total' rows
        strat = strategy_df[~strategy_df['strategy'].str.contains('Total', case=False, na=False)].copy()
        quarters = strat['quarter'].unique()
        hhi_rows = []
        for q in quarters:
            q_data = strat[strat['quarter'] == q]
            total_nav = q_data['nav'].sum()
            if total_nav > 0:
                shares = q_data['nav'] / total_nav
                hhi = (shares ** 2).sum()
                hhi_rows.append({'quarter': q, 'strategy_hhi': hhi})
        if hhi_rows:
            metrics['strategy_hhi'] = pd.DataFrame(hhi_rows)

    # --- Liquidity mismatch (portfolio liquid at 30d - investor redeemable at 30d) ---
    if not liquidity_df.empty:
        inv = liquidity_df[(liquidity_df['liquidity_type'] == 'investor_liquidity') &
                           (liquidity_df['period'] == 'At most 30 days')]
        port = liquidity_df[(liquidity_df['liquidity_type'] == 'portfolio_liquidity') &
                            (liquidity_df['period'] == 'At most 30 days')]
        fin = liquidity_df[(liquidity_df['liquidity_type'] == 'financing_liquidity') &
                           (liquidity_df['period'] == 'At most 30 days')]
        if not inv.empty and not port.empty:
            liq = inv[['quarter', 'cumulative_pct']].rename(columns={'cumulative_pct': 'investor_30d'})
            liq = liq.merge(
                port[['quarter', 'cumulative_pct']].rename(columns={'cumulative_pct': 'portfolio_30d'}),
                on='quarter', how='outer')
            if not fin.empty:
                liq = liq.merge(
                    fin[['quarter', 'cumulative_pct']].rename(columns={'cumulative_pct': 'financing_30d'}),
                    on='quarter', how='outer')
            liq['liquidity_mismatch_30d'] = liq['portfolio_30d'] - liq['investor_30d']
            if 'financing_30d' in liq.columns:
                liq['financing_gap_30d'] = liq['portfolio_30d'] - liq['financing_30d']
            metrics['liquidity_mismatch'] = liq

    return metrics


# ---------------------------------------------------------------------------
# Master parse function
# ---------------------------------------------------------------------------

def parse_all_form_pf(data_dir=None, output_dir=None):
    """Parse all Form PF data and produce processed CSVs."""
    if data_dir is None:
        data_dir = DATA_DIR
    if output_dir is None:
        output_dir = OUTPUT_DIR

    os.makedirs(output_dir, exist_ok=True)
    filepath = _find_latest_excel(data_dir)
    print(f"Parsing: {os.path.basename(filepath)}")

    # --- Parse all tables ---
    print("  Parsing fund counts (Section 1)...")
    fund_counts = parse_fund_counts(filepath)

    print("  Parsing GAV/NAV (Section 2)...")
    gav_nav = parse_gav_nav(filepath)

    print("  Parsing borrowing (Sections 2, 8)...")
    borrowing = parse_borrowing(filepath)

    print("  Parsing derivatives (Section 5)...")
    derivatives = parse_derivatives(filepath)

    print("  Parsing concentration (Section 6)...")
    concentration = parse_concentration(filepath)

    print("  Parsing strategy allocation (Section 8)...")
    strategy = parse_strategy(filepath)

    print("  Parsing leverage distributions (Section 8)...")
    leverage_dist = parse_leverage_distribution(filepath)

    print("  Parsing notional exposures (Section 8)...")
    notional = parse_notional(filepath)

    print("  Parsing liquidity (Section 8)...")
    liquidity = parse_liquidity(filepath)

    print("  Parsing fair value hierarchy (Section 2)...")
    fair_value = parse_fair_value(filepath)

    print("  Parsing geography (Section 3)...")
    geography = parse_geography(filepath)

    print("  Parsing sector (Section 10)...")
    sector = parse_sector(filepath)

    # --- Save all CSVs ---
    saves = {
        'form_pf_fund_counts.csv': fund_counts,
        'form_pf_gav_nav.csv': gav_nav,
        'form_pf_derivatives.csv': derivatives,
        'form_pf_concentration.csv': concentration,
        'form_pf_strategy.csv': strategy,
        'form_pf_leverage_dist.csv': leverage_dist,
        'form_pf_notional.csv': notional,
        'form_pf_liquidity.csv': liquidity,
        'form_pf_fair_value.csv': fair_value,
        'form_pf_geography.csv': geography,
        'form_pf_sector.csv': sector,
    }

    # Borrowing has multiple sub-tables
    if isinstance(borrowing, dict):
        if not borrowing['quarterly_pct'].empty:
            saves['form_pf_borrowing_pct.csv'] = borrowing['quarterly_pct']
        if not borrowing['monthly_detail'].empty:
            saves['form_pf_borrowing_detail.csv'] = borrowing['monthly_detail']
        if not borrowing['creditor_type'].empty:
            saves['form_pf_borrowing_creditor.csv'] = borrowing['creditor_type']

    for filename, df in saves.items():
        if df is not None and not df.empty:
            path = os.path.join(output_dir, filename)
            df.to_csv(path, index=False)
            print(f"  Saved {filename} ({len(df)} rows)")
        else:
            print(f"  Skipped {filename} (empty)")

    # --- Compute and save derived metrics ---
    print("\n  Computing derived metrics...")
    metrics = compute_form_pf_metrics(gav_nav, concentration, notional, liquidity, strategy)
    for name, mdf in metrics.items():
        if mdf is not None and not mdf.empty:
            path = os.path.join(output_dir, f'form_pf_metric_{name}.csv')
            mdf.to_csv(path, index=False)
            print(f"  Saved form_pf_metric_{name}.csv ({len(mdf)} rows)")

    # --- Summary ---
    print(f"\nDone! Parsed {os.path.basename(filepath)}")
    if not gav_nav.empty:
        hf = gav_nav[gav_nav['fund_type'] == 'Hedge Fund']
        if not hf.empty:
            latest = hf.iloc[-1]
            print(f"  Latest HF GAV: ${latest['gav']:.0f}B")
            print(f"  Latest HF NAV: ${latest['nav']:.0f}B")
            print(f"  GAV/NAV ratio:  {latest['gav_nav_ratio']:.2f}x")

    if not concentration.empty and 'nav_share' in concentration.columns:
        top10 = concentration[concentration['top_n'] == 'Top 10']
        if not top10.empty:
            latest_c = top10.iloc[-1]
            print(f"  Top-10 NAV share: {latest_c['nav_share']:.1%}")


if __name__ == '__main__':
    parse_all_form_pf()
