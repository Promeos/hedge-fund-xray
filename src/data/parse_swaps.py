"""
CFTC Weekly Swaps Report Parser

Parses weekly swap reports from CFTC containing interest rate, credit, and FX
swap notional outstanding, cleared/uncleared splits, and counterparty breakdowns.

Each file contains 52 sheets with weekly snapshots. Sheet 1 is the overview
with all asset classes. Values are in millions USD.

Data: ~600 weekly files (2013-2026).
"""

import os
import pandas as pd
import numpy as np
import openpyxl
from datetime import datetime

DATA_DIR = os.path.join(os.path.dirname(__file__), '..', '..', 'data', 'raw', 'swaps')
OUTPUT_DIR = os.path.join(os.path.dirname(__file__), '..', '..', 'data', 'processed')


def _extract_date_from_filename(filename):
    """Extract date from CFTC_Swaps_Report_MM_DD_YYYY.xlsx"""
    name = filename.replace('.xlsx', '').replace('cftc_swaps_report_', '').replace(
        'CFTC_Swaps_Report_', '')
    parts = name.split('_')
    try:
        if len(parts) >= 3:
            month, day, year = int(parts[0]), int(parts[1]), int(parts[2])
            if year < 100:
                year += 2000
            return datetime(year, month, day)
    except (ValueError, IndexError):
        pass
    return None


def parse_overview_sheet(filepath):
    """Parse Sheet 1 (overview) — IR, Credit, FX notional with weekly columns.

    Returns a DataFrame with columns: date, metric, value (millions USD).
    """
    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    except Exception:
        return pd.DataFrame()

    if '1' not in wb.sheetnames:
        wb.close()
        return pd.DataFrame()

    ws = wb['1']
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if len(rows) < 2:
        return pd.DataFrame()

    # Row 0: header with dates
    # Rows 1-9: data (IR total/cleared/uncleared, Credit total/cleared/uncleared, FX total/cleared/uncleared)
    header = rows[0]

    # Extract dates from header
    dates = []
    for val in header[1:]:
        if val is None:
            dates.append(None)
        elif isinstance(val, datetime):
            dates.append(val)
        else:
            try:
                dates.append(pd.to_datetime(val))
            except Exception:
                dates.append(None)

    # Extract metric rows
    metric_names = [
        'ir_total', 'ir_cleared', 'ir_uncleared',
        'credit_total', 'credit_cleared', 'credit_uncleared',
        'fx_total', 'fx_cleared', 'fx_uncleared',
    ]

    records = []
    for row_idx, metric in enumerate(metric_names):
        if row_idx + 1 >= len(rows):
            break
        row = rows[row_idx + 1]
        for col_idx, date in enumerate(dates):
            if date is None:
                continue
            val = row[col_idx + 1] if col_idx + 1 < len(row) else None
            if val is not None:
                records.append({
                    'date': date,
                    'metric': metric,
                    'value_millions': pd.to_numeric(val, errors='coerce'),
                })

    return pd.DataFrame(records)


def parse_all_swaps(data_dir=None, output_dir=None):
    """Parse all weekly swap reports and produce processed CSVs."""
    if data_dir is None:
        data_dir = DATA_DIR
    if output_dir is None:
        output_dir = OUTPUT_DIR

    os.makedirs(output_dir, exist_ok=True)

    files = sorted([f for f in os.listdir(data_dir) if f.endswith('.xlsx')])
    print(f"Parsing {len(files)} CFTC weekly swap reports...")

    all_records = []
    failed = 0

    for i, f in enumerate(files):
        filepath = os.path.join(data_dir, f)
        df = parse_overview_sheet(filepath)

        if df.empty:
            failed += 1
            continue

        all_records.append(df)

        if (i + 1) % 50 == 0:
            print(f"  [{i+1}/{len(files)}] processed...")

    if not all_records:
        print("No swap data parsed!")
        return

    # Combine all records and deduplicate (files contain overlapping weekly dates)
    combined = pd.concat(all_records, ignore_index=True)
    combined = combined.drop_duplicates(subset=['date', 'metric'], keep='last')
    combined = combined.sort_values(['date', 'metric'])

    # Convert to billions for consistency
    combined['value_billions'] = combined['value_millions'] / 1000

    # --- Pivot to wide format for easier analysis ---
    wide = combined.pivot_table(index='date', columns='metric',
                                values='value_billions', aggfunc='first')
    wide = wide.reset_index()
    wide.columns.name = None

    # Compute derived metrics
    if 'ir_total' in wide.columns and 'ir_cleared' in wide.columns:
        wide['ir_cleared_pct'] = wide['ir_cleared'] / wide['ir_total']
    if 'credit_total' in wide.columns and 'credit_cleared' in wide.columns:
        wide['credit_cleared_pct'] = wide['credit_cleared'] / wide['credit_total']
    if 'fx_total' in wide.columns and 'fx_cleared' in wide.columns:
        wide['fx_cleared_pct'] = wide['fx_cleared'] / wide['fx_total']

    # --- Save weekly time series ---
    wide.to_csv(os.path.join(output_dir, 'swaps_weekly.csv'), index=False)
    print(f"  Saved swaps_weekly.csv ({len(wide)} rows)")

    # --- Save long format ---
    combined.to_csv(os.path.join(output_dir, 'swaps_weekly_long.csv'), index=False)
    print(f"  Saved swaps_weekly_long.csv ({len(combined)} rows)")

    # --- Quarterly aggregation ---
    wide['quarter'] = pd.to_datetime(wide['date']).dt.to_period('Q').astype(str)
    q_agg = {}
    for col in wide.columns:
        if col in ['date', 'quarter']:
            continue
        q_agg[col] = 'last'  # Use quarter-end value
    q_agg[f'ir_total_mean'] = ('ir_total', 'mean') if 'ir_total' in wide.columns else None

    quarterly = wide.groupby('quarter').agg(
        weeks=('date', 'count'),
        **{col: (col, 'last') for col in wide.columns if col not in ['date', 'quarter']}
    ).reset_index()
    quarterly.to_csv(os.path.join(output_dir, 'swaps_quarterly.csv'), index=False)
    print(f"  Saved swaps_quarterly.csv ({len(quarterly)} rows)")

    # --- Summary ---
    print(f"\nDone! {len(files)} files parsed, {failed} failed.")
    latest = wide.iloc[-1]
    print(f"  Latest date: {latest['date']}")
    if 'ir_total' in wide.columns:
        print(f"  IR notional: ${latest['ir_total']:,.0f}B (cleared: {latest.get('ir_cleared_pct', 0):.1%})")
    if 'credit_total' in wide.columns:
        print(f"  Credit notional: ${latest['credit_total']:,.0f}B (cleared: {latest.get('credit_cleared_pct', 0):.1%})")
    if 'fx_total' in wide.columns:
        print(f"  FX notional: ${latest['fx_total']:,.0f}B (cleared: {latest.get('fx_cleared_pct', 0):.1%})")


if __name__ == '__main__':
    parse_all_swaps()
